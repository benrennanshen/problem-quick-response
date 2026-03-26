"""
Excel 导出服务
"""
import io
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class ExcelExportService:
    """
    Excel 导出服务类
    """

    # 表头定义（与前端显示顺序一致）
    DETAIL_HEADERS = [
        '诉求ID', '诉求分类', '所属院系', '是否超时', '受理单位',
        '受理时间', '完成时间', '处理时长', '标题', '学号/教工号',
        '姓名', '手机号', '联系方式', '状态', '满意度',
        '提交时间', '内容', '用户回复'
    ]

    # 汇总表头定义
    SUMMARY_HEADERS = [
        '统计项', '数值', '说明'
    ]

    # 列宽设置（单位：字符宽度）
    COLUMN_WIDTHS = {
        '诉求ID': 20,
        '诉求分类': 15,
        '所属院系': 15,
        '是否超时': 10,
        '受理单位': 20,
        '受理时间': 20,
        '完成时间': 20,
        '处理时长': 12,
        '标题': 30,
        '学号/教工号': 15,
        '姓名': 12,
        '手机号': 15,
        '联系方式': 15,
        '状态': 10,
        '满意度': 10,
        '提交时间': 20,
        '内容': 40,
        '用户回复': 30,
        '统计项': 20,
        '数值': 15,
        '说明': 30
    }

    @staticmethod
    def _create_header_style() -> Dict:
        """
        创建表头样式
        """
        return {
            'font': Font(name='微软雅黑', size=11, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    @staticmethod
    def _create_cell_style() -> Dict:
        """
        创建普通单元格样式
        """
        return {
            'font': Font(name='微软雅黑', size=10),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    @staticmethod
    def _create_number_style() -> Dict:
        """
        创建数字单元格样式
        """
        return {
            'font': Font(name='微软雅黑', size=10),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def _apply_cell_style(self, cell, style_dict: Dict):
        """
        应用样式到单元格
        """
        if 'font' in style_dict:
            cell.font = style_dict['font']
        if 'fill' in style_dict:
            cell.fill = style_dict['fill']
        if 'alignment' in style_dict:
            cell.alignment = style_dict['alignment']
        if 'border' in style_dict:
            cell.border = style_dict['border']

    def _set_column_widths(self, worksheet, headers: List[str]):
        """
        设置列宽
        """
        for col_idx, header in enumerate(headers, start=1):
            width = self.COLUMN_WIDTHS.get(header, 15)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _create_detail_sheet(self, workbook: Workbook, records: List[Dict]):
        """
        创建明细数据工作表
        """
        worksheet = workbook.active
        worksheet.title = "明细数据"

        header_style = self._create_header_style()
        cell_style = self._create_cell_style()

        # 写入表头
        for col_idx, header in enumerate(self.DETAIL_HEADERS, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            self._apply_cell_style(cell, header_style)

        # 写入数据行
        for row_idx, record in enumerate(records, start=2):
            for col_idx, header in enumerate(self.DETAIL_HEADERS, start=1):
                value = record.get(header, '')
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                self._apply_cell_style(cell, cell_style)

        # 设置列宽
        self._set_column_widths(worksheet, self.DETAIL_HEADERS)

        # 冻结首行
        worksheet.freeze_panes = 'A2'

    def _create_summary_sheet(self, workbook: Workbook, summary: Dict):
        """
        创建数据汇总工作表
        """
        worksheet = workbook.create_sheet(title="数据汇总")

        header_style = self._create_header_style()
        cell_style = self._create_cell_style()
        number_style = self._create_number_style()

        # 写入表头
        for col_idx, header in enumerate(self.SUMMARY_HEADERS, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            self._apply_cell_style(cell, header_style)

        # 汇总数据映射
        summary_rows = [
            ('总记录数', summary.get('total_count', 0), '筛选条件下的总记录数'),
            ('受理量', summary.get('accepted_count', 0), '已受理的记录数'),
            ('办结量', summary.get('completed_count', 0), '已办结的记录数'),
            ('1个工作日内办结', summary.get('completed_in_one_workday', 0), '受理后1个工作日内完成'),
            ('3个工作日内办结', summary.get('completed_in_3_days', 0), '受理后3个工作日内完成（排除1日内）'),
            ('7个工作日内办结', summary.get('completed_in_7_days', 0), '受理后7个工作日内完成（排除1日/3日内）'),
            ('超过7个工作日办结', summary.get('completed_over_7_days', 0), '受理后超过7个工作日完成'),
        ]

        # 写入汇总数据
        for row_idx, (name, value, description) in enumerate(summary_rows, start=2):
            worksheet.cell(row=row_idx, column=1, value=name)
            worksheet.cell(row=row_idx, column=2, value=value)
            worksheet.cell(row=row_idx, column=3, value=description)

            # 应用样式
            for col_idx in range(1, 4):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if col_idx == 2:
                    self._apply_cell_style(cell, number_style)
                else:
                    self._apply_cell_style(cell, cell_style)

        # 设置列宽
        self._set_column_widths(worksheet, self.SUMMARY_HEADERS)

        # 冻结首行
        worksheet.freeze_panes = 'A2'

    def export_detail_to_excel(
        self,
        records: List[Dict],
        summary: Dict,
        filename: Optional[str] = None
    ) -> bytes:
        """
        导出明细数据和汇总到 Excel

        Args:
            records: 明细记录列表
            summary: 汇总统计数据
            filename: 文件名（可选，用于日志）

        Returns:
            Excel 文件的字节内容
        """
        try:
            workbook = Workbook()

            # 创建明细数据工作表
            self._create_detail_sheet(workbook, records)

            # 创建数据汇总工作表
            self._create_summary_sheet(workbook, summary)

            # 保存到内存字节流
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            excel_bytes = output.read()
            output.close()

            logger.info(f"Excel 导出成功: {len(records)} 条记录, 文件名: {filename}")
            return excel_bytes

        except Exception as e:
            logger.error(f"Excel 导出失败: {e}", exc_info=True)
            raise RuntimeError(f"Excel 导出失败: {str(e)}")
